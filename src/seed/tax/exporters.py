# Tax Data Exporters - MISA, CSV, Excel export functionality
# Supports Vietnam accounting software formats

from decimal import Decimal
from typing import Optional, List, Dict, Any, Union
from datetime import datetime
from pathlib import Path
import csv

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .models import (
    MISAExportInput, ExportFormat, ExportResult,
    InvoiceInput, ComplianceReport,
    quantize_vnd
)
from .calculator import TaxCalculator


class MISAExporter:
    """Exports tax data in MISA format (Vietnam accounting software)"""

    # MISA standard field mappings
    JOURNAL_HEADERS = [
        'NgayChungTu', 'SoChungTu', 'LoaiChungTu', 'DienGiai',
        'TaiKhoanNo', 'TaiKhoanCo', 'SoTien', 'NgoaiTe', 'TyGia',
        'SoTienNgoaiTe', 'MaVatTu', 'TenVatTu', 'DonViTinh', 'SoLuong',
        'DonGia', 'ThueSuat', 'TienThue', 'KhoanMuc', 'BoPhan',
        'DoiTuong', 'MaDoiTuong', 'CongTrinh', 'MaCongTrinh'
    ]

    VAT_REGISTER_HEADERS = [
        'Ky', 'Ngay', 'So', 'KyHieu', 'TenNguoiBan', 'MaSoThueNguoiBan',
        'DiaChiNguoiBan', 'TenNguoiMua', 'MaSoThueNguoiMua', 'DiaChiNguoiMua',
        'TenHangHoa', 'DonViTinh', 'SoLuong', 'DonGia', 'ThanhTien',
        'ThueSuat', 'TienThue', 'GhiChu'
    ]

    LEDGER_HEADERS = [
        'Ngay', 'SoChungTu', 'DienGiai', 'TaiKhoanDoiUng',
        'No', 'Co', 'SoDuNo', 'SoDuCo', 'NgoaiTe', 'TyGia'
    ]

    TEMPLATE_CODE = "TT78"

    def __init__(self, calculator: Optional[TaxCalculator] = None):
        self.calculator = calculator or TaxCalculator()

    def export_journal(
        self,
        input_data: MISAExportInput,
        transactions: List[Dict[str, Any]]
    ) -> ExportResult:
        """Export general journal in MISA format"""
        file_path = self._get_output_path(input_data, 'journal', 'csv')

        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(self.JOURNAL_HEADERS)

            for txn in transactions:
                row = [
                    txn.get('date', ''),
                    txn.get('voucher_number', ''),
                    txn.get('voucher_type', ''),
                    txn.get('description', ''),
                    txn.get('debit_account', ''),
                    txn.get('credit_account', ''),
                    self._format_amount(txn.get('amount', 0)),
                    txn.get('currency', 'VND'),
                    self._format_amount(txn.get('exchange_rate', 1)),
                    self._format_amount(txn.get('foreign_amount', 0)),
                    txn.get('item_code', ''),
                    txn.get('item_name', ''),
                    txn.get('unit', ''),
                    self._format_amount(txn.get('quantity', 0)),
                    self._format_amount(txn.get('unit_price', 0)),
                    self._format_rate(txn.get('vat_rate', 0)),
                    self._format_amount(txn.get('vat_amount', 0)),
                    txn.get('cost_center', ''),
                    txn.get('department', ''),
                    txn.get('counterparty_type', ''),
                    txn.get('counterparty_code', ''),
                    txn.get('project', ''),
                    txn.get('project_code', ''),
                ]
                writer.writerow(row)

        return ExportResult(
            format=ExportFormat.MISA,
            file_path=str(file_path),
            record_count=len(transactions),
            file_size_bytes=file_path.stat().st_size,
            generated_at=datetime.now()
        )

    def export_vat_register(
        self,
        input_data: MISAExportInput,
        invoices: List[InvoiceInput]
    ) -> ExportResult:
        """Export VAT register (Sổ hóa đơn GTGT) in MISA format"""
        file_path = self._get_output_path(input_data, 'vat_register', 'csv')

        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(self.VAT_REGISTER_HEADERS)

            for inv in invoices:
                for item in inv.line_items:
                    row = [
                        input_data.tax_year.value,
                        inv.header.invoice_date.strftime('%d/%m/%Y'),
                        inv.header.invoice_number,
                        self.TEMPLATE_CODE,
                        inv.header.seller_name,
                        inv.header.seller_tax_code,
                        inv.header.seller_address,
                        inv.header.buyer_name,
                        inv.header.buyer_tax_code or '',
                        inv.header.buyer_address,
                        item.description,
                        item.unit,
                        self._format_amount(item.quantity),
                        self._format_amount(item.unit_price),
                        self._format_amount(item.line_total),
                        self._format_rate(item.vat_rate or Decimal('0.10')),
                        self._format_amount(item.vat_amount),
                        inv.notes or '',
                    ]
                    writer.writerow(row)

        return ExportResult(
            format=ExportFormat.MISA,
            file_path=str(file_path),
            record_count=sum(len(inv.line_items) for inv in invoices),
            file_size_bytes=file_path.stat().st_size,
            generated_at=datetime.now()
        )

    def export_ledger(
        self,
        input_data: MISAExportInput,
        entries: List[Dict[str, Any]]
    ) -> ExportResult:
        """Export general ledger in MISA format"""
        file_path = self._get_output_path(input_data, 'ledger', 'csv')

        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(self.LEDGER_HEADERS)

            running_debit = Decimal('0')
            running_credit = Decimal('0')

            for entry in entries:
                debit = Decimal(str(entry.get('debit', 0)))
                credit = Decimal(str(entry.get('credit', 0)))
                running_debit += debit
                running_credit += credit

                row = [
                    entry.get('date', ''),
                    entry.get('voucher_number', ''),
                    entry.get('description', ''),
                    entry.get('counterpart_account', ''),
                    self._format_amount(debit) if debit > 0 else '',
                    self._format_amount(credit) if credit > 0 else '',
                    self._format_amount(running_debit) if running_debit > 0 else '',
                    self._format_amount(running_credit) if running_credit > 0 else '',
                    entry.get('currency', 'VND'),
                    self._format_amount(entry.get('exchange_rate', 1)),
                ]
                writer.writerow(row)

        return ExportResult(
            format=ExportFormat.MISA,
            file_path=str(file_path),
            record_count=len(entries),
            file_size_bytes=file_path.stat().st_size,
            generated_at=datetime.now()
        )

    def _get_output_path(self, input_data: MISAExportInput, data_type: str, ext: str) -> Path:
        """Generate output file path"""
        output_dir = Path('exports') / input_data.tax_year.value / input_data.company_tax_code
        output_dir.mkdir(parents=True, exist_ok=True)

        filename = f"{data_type}_{input_data.company_tax_code}_{input_data.start_date.strftime('%Y%m%d')}_{input_data.end_date.strftime('%Y%m%d')}.{ext}"
        return output_dir / filename

    def _format_amount(self, amount: Union[Decimal, float, int, str]) -> str:
        """Format amount for MISA (no decimal for VND)"""
        if isinstance(amount, str):
            amount = Decimal(amount)
        elif isinstance(amount, (int, float)):
            amount = Decimal(str(amount))
        return f"{quantize_vnd(amount):,.0f}".replace(',', '.')

    def _format_rate(self, rate: Union[Decimal, float, int, str]) -> str:
        """Format tax rate as percentage"""
        if isinstance(rate, str):
            rate = Decimal(rate)
        elif isinstance(rate, (int, float)):
            rate = Decimal(str(rate))
        return f"{rate * 100:.0f}"


class CSVExporter:
    """Exports tax data to CSV format"""

    def __init__(self, calculator: Optional[TaxCalculator] = None):
        self.calculator = calculator or TaxCalculator()

    def export_invoices(
        self,
        invoices: List[InvoiceInput],
        file_path: Union[str, Path]
    ) -> ExportResult:
        """Export invoices to CSV"""
        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            # Header
            writer.writerow([
                'Invoice Number', 'Invoice Date', 'Template',
                'Seller Name', 'Seller Tax Code', 'Seller Address',
                'Buyer Name', 'Buyer Tax Code', 'Buyer Address',
                'Line Number', 'Description', 'Unit', 'Quantity',
                'Unit Price', 'Line Total', 'VAT Category', 'VAT Rate',
                'VAT Amount', 'Payment Method', 'Currency', 'Notes'
            ])

            for inv in invoices:
                for item in inv.line_items:
                    writer.writerow([
                        inv.header.invoice_number,
                        inv.header.invoice_date.strftime('%Y-%m-%d'),
                        inv.header.invoice_template,
                        inv.header.seller_name,
                        inv.header.seller_tax_code,
                        inv.header.seller_address,
                        inv.header.buyer_name,
                        inv.header.buyer_tax_code or '',
                        inv.header.buyer_address,
                        item.line_number,
                        item.description,
                        item.unit,
                        self._format_decimal(item.quantity),
                        self._format_decimal(item.unit_price),
                        self._format_decimal(item.line_total),
                        item.vat_category.value,
                        self._format_decimal(item.vat_rate or Decimal('0.10')),
                        self._format_decimal(item.vat_amount),
                        inv.header.payment_method,
                        inv.header.currency,
                        inv.notes or '',
                    ])

        return ExportResult(
            format=ExportFormat.CSV,
            file_path=str(path),
            record_count=sum(len(inv.line_items) for inv in invoices),
            file_size_bytes=path.stat().st_size,
            generated_at=datetime.now()
        )

    def export_tncn_payroll(
        self,
        payroll_results: List[Dict[str, Any]],
        file_path: Union[str, Path]
    ) -> ExportResult:
        """Export TNCN payroll data to CSV"""
        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow([
                'Employee ID', 'Employee Name', 'Tax Year', 'Month',
                'Gross Income', 'Personal Deduction', 'Dependent Deduction',
                'Insurance Deduction', 'Charity Deduction', 'Taxable Income',
                'Tax Payable', 'Net Income'
            ])

            for row in payroll_results:
                writer.writerow([
                    row.get('employee_id', ''),
                    row.get('employee_name', ''),
                    row.get('tax_year', ''),
                    row.get('month', ''),
                    self._format_decimal(row.get('gross_income', 0)),
                    self._format_decimal(row.get('personal_deduction', 0)),
                    self._format_decimal(row.get('dependent_deduction', 0)),
                    self._format_decimal(row.get('insurance_deduction', 0)),
                    self._format_decimal(row.get('charity_deduction', 0)),
                    self._format_decimal(row.get('taxable_income', 0)),
                    self._format_decimal(row.get('tax_payable', 0)),
                    self._format_decimal(row.get('net_income', 0)),
                ])

        return ExportResult(
            format=ExportFormat.CSV,
            file_path=str(path),
            record_count=len(payroll_results),
            file_size_bytes=path.stat().st_size,
            generated_at=datetime.now()
        )

    def export_gtgt_summary(
        self,
        vat_results: List[Dict[str, Any]],
        file_path: Union[str, Path]
    ) -> ExportResult:
        """Export GTGT (VAT) summary to CSV"""
        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow([
                'Tax Year', 'Category', 'Net Amount', 'VAT Rate',
                'VAT Amount', 'Gross Amount', 'Is Export'
            ])

            for row in vat_results:
                writer.writerow([
                    row.get('tax_year', ''),
                    row.get('category', ''),
                    self._format_decimal(row.get('net_amount', 0)),
                    self._format_decimal(row.get('vat_rate', 0)),
                    self._format_decimal(row.get('vat_amount', 0)),
                    self._format_decimal(row.get('gross_amount', 0)),
                    'Yes' if row.get('is_export', False) else 'No',
                ])

        return ExportResult(
            format=ExportFormat.CSV,
            file_path=str(path),
            record_count=len(vat_results),
            file_size_bytes=path.stat().st_size,
            generated_at=datetime.now()
        )

    def export_compliance_report(
        self,
        report: ComplianceReport,
        file_path: Union[str, Path]
    ) -> ExportResult:
        """Export compliance report to CSV"""
        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['Field', 'Value'])
            writer.writerow(['Report ID', report.report_id])
            writer.writerow(['Generated At', report.generated_at.isoformat()])
            writer.writerow(['Tax Year', report.tax_year])
            writer.writerow(['Company Tax Code', report.company_tax_code])
            writer.writerow(['Company Name', report.company_name])
            writer.writerow(['Reporting Period', report.reporting_period])
            writer.writerow(['Total Tax Liability', str(report.total_tax_liability)])

            if report.tncn_summary:
                writer.writerow([])
                writer.writerow(['TNCN Summary', ''])
                for key, value in report.tncn_summary.items():
                    if key != 'details':
                        writer.writerow([f'  {key}', str(value)])

            if report.tndn_summary:
                writer.writerow([])
                writer.writerow(['TNDN Summary', ''])
                for key, value in report.tndn_summary.items():
                    writer.writerow([f'  {key}', str(value)])

            if report.gtgt_summary:
                writer.writerow([])
                writer.writerow(['GTGT Summary', ''])
                for key, value in report.gtgt_summary.items():
                    if key != 'by_category':
                        writer.writerow([f'  {key}', str(value)])

        return ExportResult(
            format=ExportFormat.CSV,
            file_path=str(path),
            record_count=1,
            file_size_bytes=path.stat().st_size,
            generated_at=datetime.now()
        )

    def _format_decimal(self, value: Union[Decimal, float, int, str]) -> str:
        """Format decimal value"""
        if isinstance(value, str):
            value = Decimal(value)
        elif isinstance(value, (int, float)):
            value = Decimal(str(value))
        return str(value)


class ExcelExporter:
    """Exports tax data to Excel (.xlsx) format"""

    def __init__(self, calculator: Optional[TaxCalculator] = None):
        self.calculator = calculator or TaxCalculator()

    def export_invoices(
        self,
        invoices: List[InvoiceInput],
        file_path: Union[str, Path]
    ) -> ExportResult:
        """Export invoices to Excel with formatting"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoices"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        headers = [
            'Invoice Number', 'Invoice Date', 'Template',
            'Seller Name', 'Seller Tax Code', 'Seller Address',
            'Buyer Name', 'Buyer Tax Code', 'Buyer Address',
            'Line Number', 'Description', 'Unit', 'Quantity',
            'Unit Price', 'Line Total', 'VAT Category', 'VAT Rate',
            'VAT Amount', 'Payment Method', 'Currency', 'Notes'
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Write data rows
        row_idx = 2
        for inv in invoices:
            for item in inv.line_items:
                values = [
                    inv.header.invoice_number,
                    inv.header.invoice_date.strftime('%Y-%m-%d'),
                    inv.header.invoice_template,
                    inv.header.seller_name,
                    inv.header.seller_tax_code,
                    inv.header.seller_address,
                    inv.header.buyer_name,
                    inv.header.buyer_tax_code or '',
                    inv.header.buyer_address,
                    item.line_number,
                    item.description,
                    item.unit,
                    float(item.quantity),
                    float(item.unit_price),
                    float(item.line_total),
                    item.vat_category.value,
                    float(item.vat_rate or Decimal('0.10')),
                    float(item.vat_amount),
                    inv.header.payment_method,
                    inv.header.currency,
                    inv.notes or '',
                ]

                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    if col_idx in (13, 14, 15, 17, 18):  # Numeric columns
                        cell.number_format = '#,##0'

                row_idx += 1

        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(headers[col_idx - 1]))
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

        wb.save(path)

        return ExportResult(
            format=ExportFormat.EXCEL,
            file_path=str(path),
            record_count=sum(len(inv.line_items) for inv in invoices),
            file_size_bytes=path.stat().st_size,
            generated_at=datetime.now()
        )

    def export_compliance_report(
        self,
        report: ComplianceReport,
        file_path: Union[str, Path]
    ) -> ExportResult:
        """Export compliance report to Excel"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()

        # Overview sheet
        ws = wb.active
        ws.title = "Overview"
        overview_data = [
            ('Report ID', report.report_id),
            ('Generated At', report.generated_at.isoformat()),
            ('Tax Year', report.tax_year),
            ('Company Tax Code', report.company_tax_code),
            ('Company Name', report.company_name),
            ('Reporting Period', report.reporting_period),
            ('Total Tax Liability', str(report.total_tax_liability)),
        ]
        for row in overview_data:
            ws.append(list(row))

        # TNCN sheet
        if report.tncn_summary:
            ws_tncn = wb.create_sheet("TNCN")
            for key, value in report.tncn_summary.items():
                if key != 'details':
                    ws_tncn.append([key, str(value)])

        # TNDN sheet
        if report.tndn_summary:
            ws_tndn = wb.create_sheet("TNDN")
            for key, value in report.tndn_summary.items():
                ws_tndn.append([key, str(value)])

        # GTGT sheet
        if report.gtgt_summary:
            ws_gtgt = wb.create_sheet("GTGT")
            for key, value in report.gtgt_summary.items():
                if key != 'by_category':
                    ws_gtgt.append([key, str(value)])

        wb.save(path)

        return ExportResult(
            format=ExportFormat.EXCEL,
            file_path=str(path),
            record_count=1,
            file_size_bytes=path.stat().st_size,
            generated_at=datetime.now()
        )


class TaxDataExporter:
    """Facade for all tax data exporters"""

    def __init__(self, calculator: Optional[TaxCalculator] = None):
        self.calculator = calculator or TaxCalculator()
        self.misa = MISAExporter(self.calculator)
        self.csv = CSVExporter(self.calculator)
        self.excel = ExcelExporter(self.calculator)

    def export(
        self,
        format: ExportFormat,
        data_type: str,
        **kwargs: Any,
    ) -> ExportResult:
        """Export data in the specified format"""
        if format == ExportFormat.MISA:
            return self._export_misa(data_type, **kwargs)
        elif format == ExportFormat.CSV:
            return self._export_csv(data_type, **kwargs)
        elif format == ExportFormat.EXCEL:
            return self._export_excel(data_type, **kwargs)
        else:
            raise ValueError(f"Unsupported export format: {format}")

    def _export_misa(self, data_type: str, **kwargs: Any) -> ExportResult:
        input_data = kwargs['input_data']
        if data_type == 'journal':
            return self.misa.export_journal(input_data, kwargs['transactions'])
        elif data_type == 'vat_register':
            return self.misa.export_vat_register(input_data, kwargs['invoices'])
        elif data_type == 'ledger':
            return self.misa.export_ledger(input_data, kwargs['entries'])
        raise ValueError(f"Unsupported MISA data type: {data_type}")

    def _export_csv(self, data_type: str, **kwargs: Any) -> ExportResult:
        if data_type == 'invoices':
            return self.csv.export_invoices(kwargs['invoices'], kwargs['file_path'])
        elif data_type == 'tncn_payroll':
            return self.csv.export_tncn_payroll(kwargs['payroll_results'], kwargs['file_path'])
        elif data_type == 'gtgt_summary':
            return self.csv.export_gtgt_summary(kwargs['vat_results'], kwargs['file_path'])
        elif data_type == 'compliance':
            return self.csv.export_compliance_report(kwargs['report'], kwargs['file_path'])
        raise ValueError(f"Unsupported CSV data type: {data_type}")

    def _export_excel(self, data_type: str, **kwargs: Any) -> ExportResult:
        if data_type == 'invoices':
            return self.excel.export_invoices(kwargs['invoices'], kwargs['file_path'])
        elif data_type == 'compliance':
            return self.excel.export_compliance_report(kwargs['report'], kwargs['file_path'])
        raise ValueError(f"Unsupported Excel data type: {data_type}")


# Convenience functions
def export_misa_journal(
    input_data: MISAExportInput,
    transactions: List[Dict[str, Any]]
) -> ExportResult:
    """Export journal in MISA format"""
    exporter = MISAExporter()
    return exporter.export_journal(input_data, transactions)


def export_misa_vat_register(
    input_data: MISAExportInput,
    invoices: List[InvoiceInput]
) -> ExportResult:
    """Export VAT register in MISA format"""
    exporter = MISAExporter()
    return exporter.export_vat_register(input_data, invoices)


def export_invoices_csv(
    invoices: List[InvoiceInput],
    file_path: Union[str, Path]
) -> ExportResult:
    """Export invoices to CSV"""
    exporter = CSVExporter()
    return exporter.export_invoices(invoices, file_path)


def export_invoices_excel(
    invoices: List[InvoiceInput],
    file_path: Union[str, Path]
) -> ExportResult:
    """Export invoices to Excel"""
    exporter = ExcelExporter()
    return exporter.export_invoices(invoices, file_path)